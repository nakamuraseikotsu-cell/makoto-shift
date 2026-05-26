# -*- coding: utf-8 -*-
"""シフトxlsxファイル群から、指定月の実データ（予定シフト・希望休）を抽出する。

入力ファイル仕様:
  - ファイル名が「【国分寺】」「【坂下】」「【人形町】」「【東小金井】」「【武蔵小金井】」で始まる .xlsx
  - 各 xlsx は月毎にシートを持つ
  - 1シート内:
      R03 ヘッダ: 月日 / 曜日 / [スタッフ名×N] / ヘルプ / xx前半 / xx後半 / xx適正数
      R05～    : 日次データ（月日=datetime, 曜日=日本語、スタッフ列の値で出勤/休/希望休 を判定）
  - 値の意味:
      数値(1)・'早'・'遅'・'前半'・'後半'  → 出勤（planned_staff にカウント）
      '希'                                 → 希望休（leave_requests に登録）
      '公'・'有'・'特'・'欠'・'夏'         → 休暇（カウントしない）
      '自宅'・'在宅'                       → 在宅扱い
      その他の漢字（'国'・'坂下' 等）       → 他店ヘルプ（自店 planned_staff に入れない）
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import re
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import pandas as pd

from .constants import is_excluded_staff_name, drop_rows_with_excluded_staff

# ファイル名プレフィックス → エリア（院名）
SHIFT_FILE_PREFIX_AREA = {
    '【国分寺】':     '国分寺',
    '【坂下】':       '小金井坂下',
    '【人形町】':     '人形町',
    '【東小金井】':   '東小金井',
    '【武蔵小金井】': '武蔵小金井',
}

JP_WEEKDAYS = set('日月火水木金土')
FW2HW_DIGITS_R = str.maketrans('０１２３４５６７８９Ｒ', '0123456789R')

# セル値分類
WORK_TOKENS = {'早', '遅', '前半', '後半', '半', '半休', '時短'}
OFF_TOKENS = {'公', '有', '特', '欠', '夏', '産', '育'}
ABSENT_TOKENS = {'自宅', '在宅'}
LEAVE_REQUEST = '希'

# 有給休暇トークン（人事ダッシュボード用に分離して集計）
PAID_LEAVE_TOKENS = {'有'}

# 固定休（年間予定として確保された休み — 結婚式・運動会等）
# シフト xlsx 内で下記トークンを含むセルは、応援対象から完全除外する
FIXED_LEAVE_TOKENS = {'特', '結婚式', '運動会', '式', '法事', '冠婚葬祭', '記念'}

# 他店ヘルプ先 — セル値（短い漢字記号）から行き先院名へのマップ
HELP_DEST_TOKEN_MAP = {
    '国': '国分寺', '国分寺': '国分寺',
    '坂': '小金井坂下', '坂下': '小金井坂下', '小金井坂下': '小金井坂下',
    '東': '東小金井', '東小金井': '東小金井',
    '武': '武蔵小金井', '武蔵小金井': '武蔵小金井',
    '人': '人形町', '人形町': '人形町',
}


def decode_help_destination(value) -> str | None:
    """セル値から他店ヘルプ先院名を推定。識別できない場合は None。"""
    if value is None:
        return None
    s = str(value).strip().replace('　', '')
    if not s:
        return None
    # 完全一致を先に試す
    if s in HELP_DEST_TOKEN_MAP:
        return HELP_DEST_TOKEN_MAP[s]
    # 長い文字列に含まれるトークン（'坂下' を '坂' より優先する）
    for tok in ['小金井坂下', '武蔵小金井', '東小金井', '人形町', '国分寺',
                '坂下', '国', '武', '東', '坂', '人']:
        if tok in s:
            return HELP_DEST_TOKEN_MAP.get(tok)
    return None


def _norm_sheet_name(name: str) -> str:
    if name is None:
        return ''
    s = str(name).strip().replace('　', '').replace(' ', '')
    s = s.translate(FW2HW_DIGITS_R)
    # 全角ハイフン/長音→半角
    s = s.replace('－', '-').replace('～', '-').replace('〜', '-')
    return s


def match_month_sheet(sheet_names: list[str], year: int, month: int) -> str | None:
    """シート名から target year/month に対応するものを推定して返す（無ければ None）"""
    yy = year % 100
    reiwa = year - 2018
    # 完全一致候補（正規化済み比較）
    exact_candidates = {
        f'{year}-{month:02d}', f'{year}-{month}',
        f'{year}.{month:02d}', f'{year}.{month}',
        f'{year}/{month:02d}', f'{year}/{month}',
        f'{yy}-{month:02d}', f'{yy}-{month}',
        f'{yy}.{month:02d}', f'{yy}.{month}',
        f'{yy}.{month}月', f'{yy}.{month:02d}月',
        f'{yy}.{month}月分', f'{yy}.{month:02d}月分',
        f'{yy}年{month}月', f'{yy}年{month:02d}月',
        f'R{reiwa}年{month}月', f'R{reiwa}年{month:02d}月',
    }
    for sn in sheet_names:
        s = _norm_sheet_name(sn)
        if s in exact_candidates:
            return sn
        # 末尾「分」「.」などを許容する部分一致
        for cand in exact_candidates:
            if s.startswith(cand):
                return sn
    return None


def detect_area_from_filename(filename: str) -> str | None:
    for prefix, area in SHIFT_FILE_PREFIX_AREA.items():
        if filename.startswith(prefix):
            return area
    return None


def find_shift_files(folder: Path) -> list[Path]:
    """フォルダ内の シフト xlsx（プレフィックスがマップに含まれるもの）を列挙"""
    if not folder.exists():
        return []
    return sorted([
        p for p in folder.iterdir()
        if p.suffix == '.xlsx' and 'シフト' in p.name
        and detect_area_from_filename(p.name) is not None
    ])


def _cell_classify(v) -> str:
    """セル値を 'work' / 'leave_request' / 'off' / 'absent' / 'help_other' / 'empty' / 'other' に分類"""
    if v is None or v == '':
        return 'empty'
    if isinstance(v, (int, float)):
        if pd.isna(v) or v == 0:
            return 'empty'
        return 'work'
    s = str(v).strip().replace('　', '')
    if s == '':
        return 'empty'
    # 希望休（'希' or '希望休'）
    if s == LEAVE_REQUEST or s == '希望休':
        return 'leave_request'
    if s in WORK_TOKENS:
        return 'work'
    if s in OFF_TOKENS:
        return 'off'
    if s in ABSENT_TOKENS:
        return 'absent'
    # 「○○-○○」のような時刻表記もシフト勤務とみなす
    if re.match(r'^\d{1,2}[:-]\d{1,2}', s):
        return 'work'
    # その他短い文字（他店の頭文字「国」「武」「坂下」等）→ ヘルプ扱い
    if len(s) <= 4:
        return 'help_other'
    return 'other'


_EXCLUDE_STAFF_NAMES = {
    'ヘルプ', 'フルタイム', 'パート・アルバイト',
    'スタッフ数', '備考',
    '月日', '曜日',  # 念のため
}


def _find_header(rows: list) -> dict | None:
    """シフト xlsx のヘッダーをパース。新旧2フォーマット両対応。

    旧フォーマット: R03 に 月日/曜日/スタッフ名/前半/後半 が全部並ぶ
    新フォーマット: R02 に 月日/曜日、 R03 にスタッフ名/前半/後半 が並ぶ
    """
    month_day_idx = weekday_idx = None
    label_row = None
    first_idx = second_idx = optimal_idx = help_idx = None
    end_row = None

    # 先頭10行をスキャンして各マーカーの位置を見つける
    # 注意: ヘルプは新フォーマットだと R02/R03 両方にあるので、
    # end_row（スタッフ名のある行）は '前半' を基準に決める
    for r_idx in range(min(10, len(rows))):
        row = rows[r_idx] or []
        for i, v in enumerate(row):
            if v is None:
                continue
            vs = str(v).strip().replace('　', '')
            if vs == '月日' and month_day_idx is None:
                month_day_idx = i
                label_row = r_idx
            elif vs == '曜日' and weekday_idx is None:
                weekday_idx = i
            elif vs == 'ヘルプ':
                # ヘルプ列のIDXは複数行に出るが、いずれも同じ列インデックスのはず
                if help_idx is None:
                    help_idx = i
            elif vs.endswith('前半') and first_idx is None:
                first_idx = i
                end_row = r_idx       # スタッフ名のある行はここ
            elif vs.endswith('後半') and second_idx is None:
                second_idx = i
                if end_row is None:
                    end_row = r_idx
            elif vs.endswith('適正数') and optimal_idx is None:
                optimal_idx = i
                if end_row is None:
                    end_row = r_idx

    if (month_day_idx is None or weekday_idx is None
            or label_row is None or end_row is None):
        return None

    # 終端列インデックス（ヘルプ列 or 前半列の先頭）
    end_candidates = [x for x in [help_idx, first_idx] if x is not None]
    if not end_candidates:
        return None
    end_idx = min(end_candidates)

    # スタッフ名は end_row の行から拾う（新旧フォーマット共通でここ）
    staff_row = rows[end_row] or []
    staff_cols = []
    for i in range(weekday_idx + 1, end_idx):
        cell = staff_row[i] if i < len(staff_row) else None
        if cell is None:
            continue
        name = str(cell).strip().replace('　', '')
        if not name:
            continue
        if name in _EXCLUDE_STAFF_NAMES:
            continue
        # 退職・除外スタッフ（池田 等）はシート読込時点でドロップ
        if is_excluded_staff_name(name):
            continue
        # 純粋に数字だけの値はスタッフ名ではない
        try:
            float(name)
            continue
        except ValueError:
            pass
        staff_cols.append({'idx': i, 'name': name})
    if not staff_cols:
        return None

    return {
        'month_day': month_day_idx,
        'weekday': weekday_idx,
        'first': first_idx,
        'second': second_idx,
        'optimal': optimal_idx,
        'help': help_idx,
        'staff_cols': staff_cols,
        'data_start': max(label_row, end_row) + 1,
    }


def parse_shift_sheet(ws, area: str, year: int, month: int) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    cols = _find_header(rows)
    if cols is None:
        return {'planned': [], 'leave': [], 'worked': [],
                'paid_leave': [], 'help_actions': [], 'fixed_leave': []}
    planned, leave, worked = [], [], []
    paid_leave: list[dict] = []      # 有給取得記録
    help_actions: list[dict] = []    # 他店ヘルプ移動記録
    fixed_leave: list[dict] = []     # 固定休（応援不可日）
    for r_idx in range(cols['data_start'], len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        if cols['month_day'] >= len(row) or cols['weekday'] >= len(row):
            continue
        date_cell = row[cols['month_day']]
        wd_cell = row[cols['weekday']]
        if not isinstance(date_cell, datetime):
            continue
        if date_cell.year != year or date_cell.month != month:
            # 「計」「合計」等を除外しつつ、シート違いの行をスキップ
            continue
        wd = str(wd_cell or '').strip().replace('　', '')
        if wd not in JP_WEEKDAYS:
            continue
        date_str = date_cell.date().isoformat()
        work_count = 0
        for st in cols['staff_cols']:
            i = st['idx']
            v = row[i] if i < len(row) else None
            cls = _cell_classify(v)
            v_str = (str(v).strip().replace('　', '') if v is not None else '')
            if cls == 'work':
                work_count += 1
                worked.append({
                    'date': date_str,
                    'area': area,
                    'staff_name': st['name'],
                })
            elif cls == 'leave_request':
                leave.append({
                    '申請日': '',
                    'スタッフID': f'{area}_{st["name"]}',
                    'スタッフ名': st['name'],
                    '院': area,
                    '希望日': date_str,
                    '希望種別': '希望休',
                    '重要度': '中',
                    '備考': '実シフト表より抽出',
                })
            elif cls == 'off':
                # 有給休暇は別途集計（人事ダッシュボード用）
                if v_str in PAID_LEAVE_TOKENS:
                    paid_leave.append({
                        'date': date_str,
                        'area': area,
                        'staff_name': st['name'],
                        'type': '有給',
                    })
                # 固定休（年間予定の確保された休み）として応援対象から除外
                if v_str in FIXED_LEAVE_TOKENS:
                    fixed_leave.append({
                        'date': date_str,
                        'area': area,
                        'staff_name': st['name'],
                        'type': v_str,
                    })
            elif cls == 'help_other':
                dst = decode_help_destination(v_str)
                if dst and dst != area:
                    help_actions.append({
                        'date': date_str,
                        'staff_name': st['name'],
                        'src_clinic': area,
                        'dst_clinic': dst,
                        'cell_value': v_str,
                    })
        planned.append({
            'date': date_str,
            'area': area,
            'planned_staff': int(work_count),
        })
    return {
        'planned': planned, 'leave': leave, 'worked': worked,
        'paid_leave': paid_leave, 'help_actions': help_actions,
        'fixed_leave': fixed_leave,
    }


def load_real_shift_data(folder: Path, target_month: str) -> dict:
    """全シフト xlsx を走査し、target_month の予定シフト・希望休・出勤明細をDFで返す。

    ★ 印刷設定とプリンタ通信:
       openpyxl.load_workbook(read_only=True) は xlsx をストリーミング読み込みするだけで
       プリンタ API を一切呼び出さない。プリンタ通信の懸念は本関数の経路には存在しない。
       Excel アプリで該当 xlsx を開いた際の印刷設定読込を抑止したい場合は、
       本ファイル下部の sanitize_xlsx_print_settings() を別途実行すること。

    返り値:
        {
          'planned': DataFrame [date, area, planned_staff],
          'leave':   DataFrame [申請日, スタッフID, スタッフ名, 院, 希望日, 希望種別, 重要度, 備考],
          'worked':  DataFrame [date, area, staff_name]  # 日別×スタッフ名の出勤明細
          'found_areas':   set,
          'missing_areas': list,
          'files_checked': list[str],
          'matched_sheets': {area: sheet_name},
        }
    """
    year, month = map(int, target_month.split('-'))
    all_planned = []
    all_leave = []
    all_worked = []
    all_paid_leave = []
    all_help_actions = []
    all_fixed_leave = []
    found_areas = set()
    matched_sheets = {}
    files = find_shift_files(folder)

    # ★ プリンタフリーズ根本対策：読込前に各 xlsx の表示モードを 'normal' に強制
    # ★ XMLレベル（zip 直接編集）で sheetView だけを安全に変更する。
    #    openpyxl の load+save round-trip は数式の cached value を破損させるため
    #    使用しない（過去にバグあり）。
    # 失敗時（Excel で開いている / 想定外の構造など）は警告のみで処理継続
    sanitize_summary = {'ok': 0, 'noop': 0, 'locked': 0, 'error': 0}
    for path in files:
        res = force_normal_view_xml(path)
        sanitize_summary[res['status']] = sanitize_summary.get(res['status'], 0) + 1
        if res['status'] == 'locked':
            print(f'  [WARN] {path.name} は Excel で開かれているため '
                  f'view 正規化をスキップ（読込は継続）')
        elif res['status'] == 'error':
            print(f'  [WARN] {path.name} の view 正規化エラー（読込は継続）: '
                  f'{res.get("error", "")}')
        elif res['status'] == 'ok':
            print(f'  [sanitize] {path.name}: '
                  f'{res["changed"]} シートを normal ビューに変更（XML安全モード）')
    if sanitize_summary['ok'] or sanitize_summary['locked'] or sanitize_summary['error']:
        print(f'[sanitize] view 正規化サマリ: {sanitize_summary}')

    for path in files:
        area = detect_area_from_filename(path.name)
        if not area:
            continue
        try:
            # read_only=True: ストリーミング読込のためプリンタ通信は発生しない
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f'  [WARN] {path.name} の読込失敗: {e}')
            continue
        sheet_name = match_month_sheet(wb.sheetnames, year, month)
        if sheet_name is None:
            wb.close()
            continue
        ws = wb[sheet_name]
        result = parse_shift_sheet(ws, area, year, month)
        wb.close()
        if result['planned']:
            all_planned.extend(result['planned'])
            all_leave.extend(result['leave'])
            all_worked.extend(result.get('worked', []))
            all_paid_leave.extend(result.get('paid_leave', []))
            all_help_actions.extend(result.get('help_actions', []))
            all_fixed_leave.extend(result.get('fixed_leave', []))
            found_areas.add(area)
            matched_sheets[area] = sheet_name

    expected_areas = list(SHIFT_FILE_PREFIX_AREA.values())
    missing = [a for a in expected_areas if a not in found_areas]

    df_planned = (pd.DataFrame(all_planned) if all_planned
                  else pd.DataFrame(columns=['date', 'area', 'planned_staff']))
    df_leave = (pd.DataFrame(all_leave) if all_leave
                else pd.DataFrame(columns=['申請日', 'スタッフID', 'スタッフ名',
                                            '院', '希望日', '希望種別',
                                            '重要度', '備考']))
    df_worked = (pd.DataFrame(all_worked) if all_worked
                 else pd.DataFrame(columns=['date', 'area', 'staff_name']))
    df_paid_leave = (pd.DataFrame(all_paid_leave) if all_paid_leave
                     else pd.DataFrame(columns=['date', 'area',
                                                'staff_name', 'type']))
    df_help_actions = (pd.DataFrame(all_help_actions) if all_help_actions
                       else pd.DataFrame(columns=['date', 'staff_name',
                                                  'src_clinic', 'dst_clinic',
                                                  'cell_value']))
    df_fixed_leave = (pd.DataFrame(all_fixed_leave) if all_fixed_leave
                      else pd.DataFrame(columns=['date', 'area',
                                                 'staff_name', 'type']))

    # 安全網: 退職・除外スタッフ（池田 等）を念のため DataFrame 段階でも落とす
    df_leave = drop_rows_with_excluded_staff(df_leave,
                                              staff_columns=['スタッフ名',
                                                              'スタッフID'])
    df_worked = drop_rows_with_excluded_staff(df_worked,
                                               staff_columns=['staff_name'])
    df_paid_leave = drop_rows_with_excluded_staff(
        df_paid_leave, staff_columns=['staff_name'])
    df_help_actions = drop_rows_with_excluded_staff(
        df_help_actions, staff_columns=['staff_name'])
    df_fixed_leave = drop_rows_with_excluded_staff(
        df_fixed_leave, staff_columns=['staff_name'])

    return {
        'planned': df_planned,
        'leave': df_leave,
        'worked': df_worked,
        'paid_leave': df_paid_leave,
        'help_actions': df_help_actions,
        'fixed_leave': df_fixed_leave,
        'found_areas': found_areas,
        'missing_areas': missing,
        'files_checked': [p.name for p in files],
        'matched_sheets': matched_sheets,
    }


# =============================================================================
# プリンタ通信回避ユーティリティ
# -----------------------------------------------------------------------------
# Excel ファイル内に印刷領域(print_area) / 印刷タイトル / 改ページプレビュー設定が
# 残っていると、利用環境によっては Excel アプリ起動時にデフォルトプリンタへ
# 接続を試みて UI がフリーズするケースがある。
# 下記関数は openpyxl で xlsx を開いて以下を全シートで強制クリア・保存する:
#   - print_area / print_title_rows / print_title_cols
#   - page_setup / page_margins （初期値に戻す）
#   - row_breaks / col_breaks （改ページを全て削除）
#   - sheet_view.view = 'normal' （標準ビューに固定）
#   - sheet_view.zoomScalePageLayoutView / zoomScaleSheetLayoutView を None に
# read_only モードのリード処理には一切影響しない。書き戻し前提の opt-in 関数。
# =============================================================================
def sanitize_xlsx_print_settings(path: Path) -> dict:
    """xlsx の印刷・改ページ・ビュー設定を全シートでクリアして上書き保存。

    Returns: {sheets_cleaned: int, breaks_removed: int, status: 'ok'/'skip'/'error',
              path: str, error: str|None}
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() != '.xlsx':
        return {'status': 'skip', 'path': str(p), 'error': 'not xlsx',
                'sheets_cleaned': 0, 'breaks_removed': 0}
    try:
        # read_only=False: 書き換え可能モード。それでも printer 接続は openpyxl では起きない。
        wb = openpyxl.load_workbook(p, data_only=False)
    except Exception as e:
        return {'status': 'error', 'path': str(p), 'error': str(e),
                'sheets_cleaned': 0, 'breaks_removed': 0}

    sheets_cleaned = 0
    breaks_removed = 0
    try:
        for ws in wb.worksheets:
            # 印刷領域 / 印刷タイトル
            try:
                ws.print_area = None
            except Exception:
                pass
            try:
                ws.print_title_rows = None
                ws.print_title_cols = None
            except Exception:
                pass
            # 改ページ（行・列）
            try:
                rb = getattr(ws, 'row_breaks', None)
                cb = getattr(ws, 'col_breaks', None)
                if rb is not None and hasattr(rb, 'brk'):
                    breaks_removed += len(rb.brk)
                    rb.brk = []
                if cb is not None and hasattr(cb, 'brk'):
                    breaks_removed += len(cb.brk)
                    cb.brk = []
            except Exception:
                pass
            # ページ設定（フィット/印刷向き/プリンタ依存設定）を中立化
            try:
                ps = ws.page_setup
                ps.orientation = None
                ps.paperSize = None
                ps.fitToWidth = None
                ps.fitToHeight = None
                ps.scale = None
                ps.firstPageNumber = None
                ps.useFirstPageNumber = None
                ps.horizontalDpi = None
                ps.verticalDpi = None
                ps.copies = None
                ps.draft = None
                ps.cellComments = None
                ps.errors = None
            except Exception:
                pass
            # 印刷オプション（中央寄せ・グリッド線印刷など）
            try:
                po = ws.print_options
                po.horizontalCentered = None
                po.verticalCentered = None
                po.headings = None
                po.gridLines = None
                po.gridLinesSet = None
            except Exception:
                pass
            # シートビューを「標準」に固定（改ページプレビュー / ページレイアウトを解除）
            try:
                sv = ws.sheet_view
                sv.view = 'normal'
                sv.zoomScalePageLayoutView = None
                sv.zoomScaleSheetLayoutView = None
                sv.zoomScaleNormal = None
            except Exception:
                pass
            sheets_cleaned += 1
        wb.save(p)
        wb.close()
        return {'status': 'ok', 'path': str(p), 'error': None,
                'sheets_cleaned': sheets_cleaned,
                'breaks_removed': breaks_removed}
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {'status': 'error', 'path': str(p), 'error': str(e),
                'sheets_cleaned': sheets_cleaned,
                'breaks_removed': breaks_removed}


def sanitize_all_shift_xlsx(folder: Path) -> list[dict]:
    """folder 内の シフト xlsx 全件に sanitize_xlsx_print_settings を適用"""
    results = []
    for path in find_shift_files(Path(folder)):
        results.append(sanitize_xlsx_print_settings(path))
    return results


# =============================================================================
# プリンタフリーズ根本対策：表示モードを「標準ビュー」に強制
# -----------------------------------------------------------------------------
# Excel が「ページレイアウト」「改ページプレビュー」モードでファイルを開くと、
# プリンタドライバへ印刷可能領域を問い合わせるためフリーズすることがある。
# このユーティリティは sheet_view.view を 'normal' に強制するのみの最小侵襲版で、
# 印刷領域や改ページの設定は触らない（=ユーザの印刷設定を尊重する）。
# load_real_shift_data の入口でも自動的に走るため、Streamlit からの読込時に
# Drive 上の各 xlsx は常に「標準ビュー」に整えられる。
# =============================================================================
def apply_normal_view_to_workbook(wb) -> int:
    """workbook 内の全シートの sheet_view.view を 'normal' に強制（保存はしない）

    新規生成した openpyxl Workbook を保存する直前など、書き込み側からも
    再利用できる公開ヘルパ。

    Returns: 変更を実施したシート数
    """
    changed = 0
    for ws in wb.worksheets:
        try:
            sv = ws.sheet_view
            if getattr(sv, 'view', None) != 'normal':
                sv.view = 'normal'
                changed += 1
        except Exception:
            # 何らかの理由で sheet_view が触れないシートは無視
            pass
    return changed


def force_normal_view(path: Path) -> dict:
    """指定 xlsx の全シートの表示モードを 'normal' に強制して上書き保存。

    冪等: 既に normal の場合は no-op で保存もスキップ。
    Excel で同ファイルを開いていてロックされている場合は status='locked' を返す。

    Returns: {'status': 'ok'/'noop'/'locked'/'error'/'skip',
              'path': str, 'changed': int, 'error': str|None}
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() != '.xlsx':
        return {'status': 'skip', 'path': str(p), 'changed': 0,
                'error': 'not xlsx'}
    try:
        # data_only=False, read_only=False で開いて書き戻し可能にする
        wb = openpyxl.load_workbook(p, data_only=False)
    except Exception as e:
        return {'status': 'error', 'path': str(p), 'changed': 0,
                'error': f'load failed: {e}'}

    try:
        changed = apply_normal_view_to_workbook(wb)
        if changed == 0:
            # 既に全シート normal なので保存不要（disk I/O 節約）
            wb.close()
            return {'status': 'noop', 'path': str(p), 'changed': 0,
                    'error': None}
        wb.save(p)
        wb.close()
        return {'status': 'ok', 'path': str(p), 'changed': changed,
                'error': None}
    except PermissionError as e:
        # Excel で開いていてロックされている等
        try:
            wb.close()
        except Exception:
            pass
        return {'status': 'locked', 'path': str(p), 'changed': 0,
                'error': str(e)}
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {'status': 'error', 'path': str(p), 'changed': 0,
                'error': str(e)}


def force_normal_view_all(folder: Path) -> list[dict]:
    """folder 内の全シフト xlsx に force_normal_view を実行"""
    return [force_normal_view(p) for p in find_shift_files(Path(folder))]


# =============================================================================
# XMLレベルでの安全な sheet_view 強制（formula 破損リスクなし）
# -----------------------------------------------------------------------------
# 重要: openpyxl の load_workbook+save の round-trip は、xlsx 内の数式セルの
#       「cached value」を消去してしまう。日付・曜日が =A6+1 や =TEXT(A6,"aaa")
#       のような数式で生成されているシートを openpyxl で読み書きすると、
#       data_only=True で読込んだ際に None になり、後続の解析が壊れる。
#
# 本ユーティリティは xlsx (= zip) を直接解凍し、各 worksheet XML の
#   <sheetView ... view="pageBreakPreview"> 属性を view="normal" に書き換える
# のみを行う。セル値・数式・印刷設定・書式は一切触らない。
# =============================================================================
_SHEET_VIEW_RE = re.compile(
    rb'(<sheetView\b[^>]*\bview=")(pageBreakPreview|pageLayout)(")'
)


def force_normal_view_xml(path: Path) -> dict:
    """xlsx の sheetView/view 属性を XML レベルで 'normal' に書き換え。

    安全性:
      - xlsx を zip として解凍 → sheetN.xml のみ正規表現で書き換え → zip 再圧縮
      - 数式・cached value・cell 書式・印刷設定・グラフ・コメント等は一切変更しない
      - 元 zip のエントリ順序とメタデータは保持

    Returns: {'status': 'ok'/'noop'/'locked'/'error'/'skip',
              'path': str, 'changed': int, 'error': str|None}
    """
    p = Path(path)
    if not p.exists() or p.suffix.lower() != '.xlsx':
        return {'status': 'skip', 'path': str(p), 'changed': 0,
                'error': 'not xlsx'}

    # 書込み権限チェック（先に試して locked を素早く判定）
    try:
        with open(p, 'r+b'):
            pass
    except PermissionError as e:
        return {'status': 'locked', 'path': str(p), 'changed': 0,
                'error': str(e)}
    except Exception as e:
        return {'status': 'error', 'path': str(p), 'changed': 0,
                'error': f'open check failed: {e}'}

    tmp_out = None
    try:
        # 一時ファイルに新しい zip を構築
        tmp_fd = tempfile.NamedTemporaryFile(
            delete=False, suffix='.xlsx',
            dir=p.parent,  # 同じドライブで move を高速化
        )
        tmp_out = Path(tmp_fd.name)
        tmp_fd.close()

        changed = 0
        with zipfile.ZipFile(p, 'r') as zin:
            with zipfile.ZipFile(tmp_out, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    # xl/worksheets/sheet*.xml の sheetView だけ書き換え
                    if (item.filename.startswith('xl/worksheets/sheet')
                            and item.filename.endswith('.xml')):
                        new_data, n = _SHEET_VIEW_RE.subn(
                            rb'\1normal\3', data
                        )
                        if n > 0:
                            data = new_data
                            changed += n
                    # ZipInfo メタデータ（タイムスタンプ等）は item を流用
                    zout.writestr(item, data)

        if changed == 0:
            # 変更不要 → 一時ファイル削除して noop
            try:
                tmp_out.unlink()
            except Exception:
                pass
            return {'status': 'noop', 'path': str(p), 'changed': 0,
                    'error': None}

        # 元ファイルを置き換え（atomic に近い）
        shutil.move(str(tmp_out), str(p))
        return {'status': 'ok', 'path': str(p), 'changed': changed,
                'error': None}
    except PermissionError as e:
        if tmp_out and tmp_out.exists():
            try: tmp_out.unlink()
            except Exception: pass
        return {'status': 'locked', 'path': str(p), 'changed': 0,
                'error': str(e)}
    except Exception as e:
        if tmp_out and tmp_out.exists():
            try: tmp_out.unlink()
            except Exception: pass
        return {'status': 'error', 'path': str(p), 'changed': 0,
                'error': str(e)}


def force_normal_view_xml_all(folder: Path) -> list[dict]:
    """folder 内の全シフト xlsx に force_normal_view_xml を実行"""
    return [force_normal_view_xml(p) for p in find_shift_files(Path(folder))]
