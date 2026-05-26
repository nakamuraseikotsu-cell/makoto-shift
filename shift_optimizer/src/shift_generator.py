# -*- coding: utf-8 -*-
"""シフト自動最適化（PuLP使用）

入力:
  - 対象月 YYYY-MM
  - 実シフトxlsx からスタッフ名簿と希望休「希」を抽出

制約:
  1) 各スタッフ 月間出勤 ≤ 22日 （目標22日）
  2) 7日連続出勤の禁止（最大 6連勤）
  3) 「希」とマークされた日は必ず公休
  4) 院別 × 曜日別 の必要人数を満たすこと（充足できない分は不足ログ）
  5) 1日に複数院掛け持ちは不可（1日1院）
  6) 中央線4院（国分寺/小金井坂下/東小金井/武蔵小金井）は相互応援可
  7) 人形町は独立、他院との行き来なし

出力:
  - assignments: 配置DataFrame
  - shortages : 不足DataFrame
"""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import calendar
import openpyxl
import pandas as pd
import pulp

from .real_data import (
    SHIFT_FILE_PREFIX_AREA, detect_area_from_filename, find_shift_files,
    match_month_sheet, _find_header, parse_shift_sheet,
)


# =====================================================================
# 設定値（ユーザー指定）
# =====================================================================
REQUIRED_BY_DOW = {
    '国分寺':     {'月': 3, '火': 2, '水': 2, '木': 3, '金': 3, '土': 4, '日': 4},
    '小金井坂下': {'月': 2, '火': 2, '水': 1, '木': 1, '金': 2, '土': 3, '日': 2},
    '東小金井':   {'月': 2, '火': 2, '水': 2, '木': 2, '金': 2, '土': 3, '日': 3},
    '武蔵小金井': {'月': 3, '火': 3, '水': 2, '木': 3, '金': 3, '土': 4, '日': 4},
    '人形町':     {'月': 2, '火': 2, '水': 2, '木': 2, '金': 2, '土': 1, '日': 2},
}

CHUO_CLINICS = ['国分寺', '小金井坂下', '東小金井', '武蔵小金井']
NINGYO = '人形町'

# 表示用：院 → セル記号（応援表記）
CLINIC_TO_CELL = {
    '国分寺': '国', '小金井坂下': '坂', '東小金井': '東', '武蔵小金井': '武',
    '人形町': '人',
}

WD_JP = {0: '月', 1: '火', 2: '水', 3: '木', 4: '金', 5: '土', 6: '日'}

WORKING_DAYS_TARGET = 22
MAX_CONSECUTIVE = 6


# =====================================================================
# 名簿 & 希望休 抽出
# =====================================================================
def extract_rosters_and_leave(folder: Path, target_month: str) -> dict:
    """シフトxlsx から (院 → スタッフ名リスト) と (院, 氏名, 日付) の希望休集合 を返す"""
    year, month = map(int, target_month.split('-'))
    rosters: dict[str, list[str]] = {}
    leave_set: set[tuple[str, str, str]] = set()
    matched: list[str] = []
    missing: list[str] = []
    files = find_shift_files(folder)
    for path in files:
        area = detect_area_from_filename(path.name)
        if not area:
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = match_month_sheet(wb.sheetnames, year, month)
        if sheet is None:
            wb.close()
            missing.append(area)
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        cols = _find_header(rows)
        if not cols:
            wb.close()
            missing.append(area)
            continue
        rosters[area] = [s['name'] for s in cols['staff_cols']]
        # 希望休抽出
        result = parse_shift_sheet(ws, area, year, month)
        for lv in result['leave']:
            leave_set.add((area, lv['スタッフ名'], lv['希望日']))
        matched.append(area)
        wb.close()
    return {
        'rosters': rosters,
        'leave': leave_set,
        'matched_areas': matched,
        'missing_areas': missing,
    }


# =====================================================================
# 最適化本体
# =====================================================================
def optimize_shift(target_month: str,
                    crm_folder: Path,
                    time_limit: int = 90) -> dict:
    year, month = map(int, target_month.split('-'))
    last_day = calendar.monthrange(year, month)[1]
    days = [datetime(year, month, d) for d in range(1, last_day + 1)]
    n_days = len(days)
    date_strs = [d.date().isoformat() for d in days]
    day_dow = {d.date().isoformat(): WD_JP[d.weekday()] for d in days}

    # ---------- 名簿 / 希望休 ----------
    info = extract_rosters_and_leave(crm_folder, target_month)
    if not info['rosters']:
        raise RuntimeError(
            f'{target_month} のシフト名簿が取得できる院がありません。'
            f'各シフトxlsxに対象月のシートを作成してください。'
        )

    # 中央線スタッフを union（同一名は1名扱い）
    chuo_home: dict[str, str] = {}
    for c in CHUO_CLINICS:
        for s in info['rosters'].get(c, []):
            if s not in chuo_home:
                chuo_home[s] = c
    chuo_names = list(chuo_home.keys())

    ningyo_staff = list(info['rosters'].get(NINGYO, []))

    # 希望休セット
    chuo_leave: set[tuple[str, str]] = set()
    ningyo_leave: set[tuple[str, str]] = set()
    for (clinic, staff, date_str) in info['leave']:
        if clinic in CHUO_CLINICS and staff in chuo_home:
            chuo_leave.add((staff, date_str))
        elif clinic == NINGYO and staff in ningyo_staff:
            ningyo_leave.add((staff, date_str))

    print(f'[ShiftGen] 中央線スタッフ {len(chuo_names)}名, 人形町スタッフ {len(ningyo_staff)}名')
    print(f'[ShiftGen] 希望休 中央線 {len(chuo_leave)}件, 人形町 {len(ningyo_leave)}件')

    # ---------- LP ----------
    prob = pulp.LpProblem('shift_opt', pulp.LpMinimize)

    # 変数: 中央線 x[s,d,c], 人形町 x[s,d]
    x_chuo = {(s, d, c): pulp.LpVariable(f'xC_{i}_{d}_{c}', cat='Binary')
              for i, s in enumerate(chuo_names) for d in date_strs for c in CHUO_CLINICS}
    x_ning = {(s, d): pulp.LpVariable(f'xN_{i}_{d}', cat='Binary')
              for i, s in enumerate(ningyo_staff) for d in date_strs}

    # 不足スラック
    u_chuo = {(d, c): pulp.LpVariable(f'uC_{d}_{c}', lowBound=0)
              for d in date_strs for c in CHUO_CLINICS}
    u_ning = {d: pulp.LpVariable(f'uN_{d}', lowBound=0) for d in date_strs}

    # 出勤日数の偏差（22 から下振れたら罰）
    dev_low_chuo = {s: pulp.LpVariable(f'dlC_{i}', lowBound=0)
                    for i, s in enumerate(chuo_names)}
    dev_low_ning = {s: pulp.LpVariable(f'dlN_{i}', lowBound=0)
                    for i, s in enumerate(ningyo_staff)}

    # ===== 制約 =====
    # 1) 1日1院（中央線）
    for s in chuo_names:
        for d in date_strs:
            prob += pulp.lpSum(x_chuo[(s, d, c)] for c in CHUO_CLINICS) <= 1

    # 2a) 月間出勤 ≤ 22 + 偏差変数
    for s in chuo_names:
        total = pulp.lpSum(x_chuo[(s, d, c)] for d in date_strs for c in CHUO_CLINICS)
        prob += total <= WORKING_DAYS_TARGET
        # total + dev_low = 22 → 22 を下回ったら dev_low > 0
        prob += total + dev_low_chuo[s] >= WORKING_DAYS_TARGET
    for s in ningyo_staff:
        total = pulp.lpSum(x_ning[(s, d)] for d in date_strs)
        prob += total <= WORKING_DAYS_TARGET
        prob += total + dev_low_ning[s] >= WORKING_DAYS_TARGET

    # 3) 希望休: 出勤させない
    for s in chuo_names:
        for d in date_strs:
            if (s, d) in chuo_leave:
                prob += pulp.lpSum(x_chuo[(s, d, c)] for c in CHUO_CLINICS) == 0
    for s in ningyo_staff:
        for d in date_strs:
            if (s, d) in ningyo_leave:
                prob += x_ning[(s, d)] == 0

    # 4) 必要人数 + 不足スラック
    for d in date_strs:
        dow = day_dow[d]
        for c in CHUO_CLINICS:
            req = REQUIRED_BY_DOW[c][dow]
            prob += (pulp.lpSum(x_chuo[(s, d, c)] for s in chuo_names)
                     + u_chuo[(d, c)] >= req)
        req_n = REQUIRED_BY_DOW[NINGYO][dow]
        prob += (pulp.lpSum(x_ning[(s, d)] for s in ningyo_staff)
                 + u_ning[d] >= req_n)

    # 5) 7日窓 ≤ 6 連勤
    for s in chuo_names:
        for start in range(n_days - 6):
            window = date_strs[start:start + 7]
            prob += pulp.lpSum(
                x_chuo[(s, dd, c)] for dd in window for c in CHUO_CLINICS
            ) <= MAX_CONSECUTIVE
    for s in ningyo_staff:
        for start in range(n_days - 6):
            window = date_strs[start:start + 7]
            prob += pulp.lpSum(x_ning[(s, dd)] for dd in window) <= MAX_CONSECUTIVE

    # 6) 小金井坂下 オープン作業リスク回避
    # ----------------------------------------------------------------
    # 「山本」が休みの日は、「稲田」1人だけでのオープン（特に朝の開店時間帯）
    # にならないよう、稲田が小金井坂下に入る場合は他スタッフ（ヘルプ含む）を
    # 必ず最低1名 同時配置する制約。
    #   ∀ d, (山本 が 坂下 にいない) ⇒
    #         (稲田_坂下 == 1 のとき、他スタッフ_坂下 ≥ 1)
    # 線形化:
    #   x[稲田,d,坂下] - sum(x[s,d,坂下] for s != 稲田,山本) ≤ 0
    #   ただし「その日 山本が出ない」ことを示すバイナリ y_d を導入して
    #   日ごとに切り替えるとモデルが複雑になるため、ここでは保守側に倒して
    #   「稲田が坂下に居る日は 必ず誰か他のスタッフが同店に居る」を
    #   山本休みの条件と切り離して常時要求する（最悪でも単独配置を防ぐ）。
    # ----------------------------------------------------------------
    SAKASHITA = '小金井坂下'
    inada_names = [s for s in chuo_names if '稲田' in s]
    yamamoto_names = [s for s in chuo_names if '山本' in s]
    if inada_names and SAKASHITA in CHUO_CLINICS:
        for inada in inada_names:
            others = [s for s in chuo_names if s != inada]
            for d in date_strs:
                # 稲田が坂下にいるとき、他のスタッフが少なくとも1名同店に居る
                prob += (
                    pulp.lpSum(x_chuo[(s, d, SAKASHITA)] for s in others)
                    >= x_chuo[(inada, d, SAKASHITA)]
                )
        print(f'[ShiftGen] リスク回避制約: 稲田 ({inada_names}) 単独オープン禁止 '
              f'(山本 {yamamoto_names})')

    # ===== 目的関数 =====
    # 1) 不足を最大の重みで最小化
    # 2) 出勤日数 22 への下振れを罰
    # 3) ホーム院出勤を促進（少しでも）
    obj_under = pulp.lpSum(u_chuo[k] for k in u_chuo) + pulp.lpSum(u_ning[k] for k in u_ning)
    obj_devlow = (pulp.lpSum(dev_low_chuo[s] for s in chuo_names)
                  + pulp.lpSum(dev_low_ning[s] for s in ningyo_staff))
    obj_home = -pulp.lpSum(
        x_chuo[(s, d, chuo_home[s])] for s in chuo_names for d in date_strs
    )
    prob += 1000 * obj_under + 10 * obj_devlow + 1 * obj_home

    # ===== 解く =====
    print(f'[ShiftGen] 最適化開始 (vars≈{len(x_chuo)+len(x_ning):,}, '
          f'time_limit={time_limit}s)')
    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=time_limit)
    status_code = prob.solve(solver)
    status_str = pulp.LpStatus[status_code]
    print(f'[ShiftGen] 終了ステータス: {status_str}')

    # ===== 結果整形 =====
    def val(v):
        x = v.value()
        return 0 if x is None else int(round(x))

    assignments = []
    shortages = []
    for d in date_strs:
        dow = day_dow[d]
        # ---- 中央線 ----
        clinic_assigned = {c: [] for c in CHUO_CLINICS}
        for s in chuo_names:
            assigned_to = None
            for c in CHUO_CLINICS:
                if val(x_chuo[(s, d, c)]) == 1:
                    assigned_to = c
                    break
            home = chuo_home[s]
            if assigned_to is None:
                if (s, d) in chuo_leave:
                    stat = '希望休'; cell = '希'
                else:
                    stat = '公休'; cell = '公'
            elif assigned_to == home:
                stat = '自院出勤'; cell = '1'
                clinic_assigned[assigned_to].append(s)
            else:
                stat = '応援出勤'; cell = CLINIC_TO_CELL.get(assigned_to, '応')
                clinic_assigned[assigned_to].append(s)
            assignments.append({
                'date': d, 'day_of_week': dow, 'staff_name': s,
                'home_clinic': home, 'assigned_clinic': assigned_to or '',
                'status': stat, 'cell': cell,
            })
        for c in CHUO_CLINICS:
            req = REQUIRED_BY_DOW[c][dow]
            asn = len(clinic_assigned[c])
            if req - asn > 0:
                shortages.append({
                    'date': d, 'day_of_week': dow, 'clinic': c,
                    'required': req, 'assigned': asn, 'gap': req - asn,
                })

        # ---- 人形町 ----
        ningyo_today = []
        for s in ningyo_staff:
            if val(x_ning[(s, d)]) == 1:
                ningyo_today.append(s)
                stat = '自院出勤'; cell = '1'; ac = NINGYO
            elif (s, d) in ningyo_leave:
                stat = '希望休'; cell = '希'; ac = ''
            else:
                stat = '公休'; cell = '公'; ac = ''
            assignments.append({
                'date': d, 'day_of_week': dow, 'staff_name': s,
                'home_clinic': NINGYO, 'assigned_clinic': ac,
                'status': stat, 'cell': cell,
            })
        req_n = REQUIRED_BY_DOW[NINGYO][dow]
        if req_n - len(ningyo_today) > 0:
            shortages.append({
                'date': d, 'day_of_week': dow, 'clinic': NINGYO,
                'required': req_n, 'assigned': len(ningyo_today),
                'gap': req_n - len(ningyo_today),
            })

    df_assign = pd.DataFrame(assignments)
    df_short = pd.DataFrame(shortages,
                            columns=['date', 'day_of_week', 'clinic',
                                     'required', 'assigned', 'gap'])

    # 出勤日数集計
    work_days = (df_assign[df_assign['status'].isin(['自院出勤', '応援出勤'])]
                 .groupby(['staff_name', 'home_clinic'])
                 .size().reset_index(name='work_days'))

    return {
        'status': status_str,
        'assignments': df_assign,
        'shortages': df_short,
        'work_days': work_days,
        'chuo_staff': chuo_home,
        'ningyo_staff': ningyo_staff,
        'leave_count': len(info['leave']),
        'missing_areas': info['missing_areas'],
        'matched_areas': info['matched_areas'],
    }
